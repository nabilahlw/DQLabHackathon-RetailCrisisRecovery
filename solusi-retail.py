import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from mlxtend.frequent_patterns import apriori, association_rules

# ── 1. LOAD & PREP DATA ────────────────────────────────────────────────────────
try:
    raw = pd.read_excel('data_penjualan.xlsx', sheet_name='Transaksi')
except Exception as e:
    print(f"Error: {e}"); exit()

raw['Tanggal'] = pd.to_datetime(raw['tgl_transaksi'])

# Agregasi harian per produk
agg = (raw
    .groupby(['tgl_transaksi', 'kode_produk', 'nama_produk'])['total_nilai']
    .sum().reset_index()
    .sort_values(['kode_produk', 'tgl_transaksi']))

# ── 2. MOVING AVERAGE & TREN ───────────────────────────────────────────────────
WINDOW = 3
agg['MA'] = agg.groupby('kode_produk')['total_nilai'].transform(
    lambda x: x.rolling(WINDOW).mean())

agg['Is_Rising'] = agg.groupby('kode_produk')['MA'].diff() > 0
agg['Trend_Session'] = (
    (agg['Is_Rising'] != agg.groupby('kode_produk')['Is_Rising'].shift())
    .groupby(agg['kode_produk']).cumsum())

def count_consec(g):
    g['Consec'] = g.groupby('Trend_Session').cumcount() + 1
    g.loc[g['Is_Rising'] == False, 'Consec'] = 0
    return g

agg = agg.groupby('kode_produk', group_keys=False).apply(count_consec)

# ── 3. NORMALISASI BASE 100 ────────────────────────────────────────────────────
def norm_base100(g):
    first = g['MA'].dropna().iloc[0] if g['MA'].dropna().size else 1
    g['Normalized'] = (g['MA'] / first) * 100
    return g

agg = agg.groupby('kode_produk', group_keys=False).apply(norm_base100)

# ── 4. GROWTH % & FILTER TREN ≥ 12 HARI ──────────────────────────────────────
MIN_DAYS = 12

growth = (agg[agg['Is_Rising']]
    .groupby(['kode_produk', 'nama_produk', 'Trend_Session'])
    .agg(Growth_Pct=('MA', lambda x: (x.iloc[-1] / x.iloc[0] - 1) * 100),
         Max_Consec=('Consec', 'max'))
    .reset_index())

top_growth = (growth[growth['Max_Consec'] >= MIN_DAYS]
    .groupby(['kode_produk', 'nama_produk'])['Growth_Pct'].max()
    .reset_index())

total_sales = raw.groupby('kode_produk')['total_nilai'].sum().reset_index()
report = (pd.merge(top_growth, total_sales, on='kode_produk')
    .sort_values('Growth_Pct', ascending=False))

# Tampilkan hasil
print(f"\nREKAPITULASI PRODUK (MINIMAL TREN {MIN_DAYS} HARI)")
print("-" * 90)
disp = report.copy()
disp.columns = ['Kode Produk', 'Nama Produk', 'Growth %', 'Total Penjualan']
disp['Growth %'] = disp['Growth %'].round(2)
disp['Total Penjualan'] = disp['Total Penjualan'].astype(int)
print(disp.to_string(index=False))

# ── 5. APRIORI / PRODUCT PACKAGING ────────────────────────────────────────────
print("\nMEMPROSES APRIORI ANALYSIS...")

# Buat basket matrix biner per struk
basket = (raw.groupby(['nomor_struk', 'nama_produk'])['jumlah_terjual']
    .sum().unstack(fill_value=0) > 0).astype(int)

freq_items = apriori(basket, min_support=0.01, use_colnames=True)
rules = association_rules(freq_items, metric='lift', min_threshold=1)

# Filter: minimal satu item adalah rising star, lift ≥ 2
rising_names = set(report['nama_produk'])
has_rising = lambda iset: any(i in rising_names for i in iset)
rules = rules[
    rules['antecedents'].apply(has_rising) |
    rules['consequents'].apply(has_rising)
]
rules = rules[rules['lift'] >= 2].sort_values(
    ['lift', 'support', 'confidence'], ascending=False)

n_trx = raw['nomor_struk'].nunique()
rules['Jumlah_Transaksi_Rule'] = (rules['support'] * n_trx).round(0).astype(int)

pkg = pd.DataFrame({
    'Jika Membeli':  rules['antecedents'].apply(lambda x: ', '.join(sorted(x, reverse=True))),
    'Maka Membeli':  rules['consequents'].apply(lambda x: ', '.join(sorted(x, reverse=True))),
    'Jumlah Invoice': rules['Jumlah_Transaksi_Rule'],
    'Support':       rules['support'].round(2),
    'Confidence':    rules['confidence'].round(2),
    'Lift':          rules['lift'].round(2),
}).sort_values(['Lift', 'Support', 'Confidence'], ascending=False)

print("\nTop 10 Packaging Recommendation:")
print(pkg.head(100))

# ── 6. EXPORT KE EXCEL ────────────────────────────────────────────────────────
OUT_FILE   = 'retail-insight.xlsx'
SHEET_RISE = 'Rising Star'
SHEET_PKG  = 'Potential Packaging'

writer_args = dict(engine='openpyxl', mode='a', if_sheet_exists='replace') \
    if os.path.exists(OUT_FILE) else dict(engine='openpyxl')

with pd.ExcelWriter(OUT_FILE, **writer_args) as w:
    disp.to_excel(w, sheet_name=SHEET_RISE, index=False)
    pkg.to_excel(w, sheet_name=SHEET_PKG, index=False)

# ── 7. FORMAT EXCEL ───────────────────────────────────────────────────────────
wb = load_workbook(OUT_FILE)
ws = wb[SHEET_RISE]

for cell in ws[1]:                         # Bold header
    cell.font = Font(bold=True)
ws.freeze_panes = 'A2'

for col in ws.columns:                     # Auto-width kolom
    ltr = get_column_letter(col[0].column)
    ws.column_dimensions[ltr].width = max(len(str(c.value or '')) for c in col) + 3

for row in ws.iter_rows(min_row=2):        # Format angka
    row[2].number_format = '0.00'
    row[3].number_format = '#,##0'

wb.save(OUT_FILE)
print(f"\nSheet '{SHEET_RISE}' berhasil dibuat.")

# ── 8. VISUALISASI: INDEKS PERTUMBUHAN (BASE 100) ─────────────────────────────
PALETTE = ['#FFD700','#C0C0C0','#CD7F32','#2ecc71',
           '#3498db','#9b59b6','#e74c3c','#34495e']
GREY3   = ['#B0B0B0','#909090','#707070']
DEFAULT_CLR = '#95a5a6'

sorted_rpt = report.sort_values('Growth_Pct', ascending=False)
clr_map  = {r.kode_produk: PALETTE[i] if i < len(PALETTE) else DEFAULT_CLR
            for i, r in enumerate(sorted_rpt.itertuples())}
rank_map = {r.kode_produk: i + 1 for i, r in enumerate(sorted_rpt.itertuples())}

top3 = (raw.groupby(['kode_produk','nama_produk'])['total_nilai']
    .sum().reset_index()
    .sort_values('total_nilai', ascending=False).head(3))
top3_df = agg[agg['kode_produk'].isin(top3['kode_produk'])]
plot_df  = agg[agg['kode_produk'].isin(report['kode_produk'])]

font_t = dict(family='sans-serif', color='black', weight='bold', size=16)
font_l = dict(family='sans-serif', weight='normal', size=12)

def add_top3(ax, y_col):
    """Plot garis Top 3 Sales dengan warna abu-abu."""
    for i, (kode, grp) in enumerate(top3_df.groupby('kode_produk')):
        ax.plot(grp['tgl_transaksi'], grp[y_col],
                linestyle='--', linewidth=2, marker='o', markersize=3,
                color=GREY3[i] if i < 3 else '#808080', alpha=0.7,
                label=f"Top Sales: {grp['nama_produk'].iloc[0]}")

def add_rising(ax, y_col):
    """Plot garis Rising Star dengan warna & rank masing-masing."""
    for kode, grp in plot_df.groupby('kode_produk'):
        ax.plot(grp['tgl_transaksi'], grp[y_col],
                marker='o', markersize=4, linewidth=2.5,
                color=clr_map.get(kode, DEFAULT_CLR),
                label=f"Rank {rank_map.get(kode,'?')}: {grp['nama_produk'].iloc[0]}")

def sort_legend(ax):
    """Pisah & urutkan legend: Top Sales dulu, lalu Rising Star by rank."""
    handles, labels = ax.get_legend_handles_labels()
    tops    = [(h,l) for h,l in zip(handles,labels) if l.startswith('Top Sales')]
    rising  = sorted([(h,l) for h,l in zip(handles,labels) if not l.startswith('Top Sales')],
                     key=lambda x: int(x[1].split(':')[0].split()[1]))
    merged  = tops + rising
    ax.legend([x[0] for x in merged], [x[1] for x in merged],
               title="Kategori Produk", title_fontsize=12, fontsize=10,
               bbox_to_anchor=(1.02,1), loc='upper left',
               borderaxespad=0, frameon=True, shadow=True)

if not plot_df.empty:
    # --- Chart 1: Indeks Base 100 ---
    fig, ax = plt.subplots(figsize=(15, 8), dpi=100)
    add_top3(ax, 'Normalized')
    add_rising(ax, 'Normalized')
    ax.set_title('ANALISIS PERTUMBUHAN RELATIF PRODUK RISING STAR\n'
                 '(Dengan Benchmark Top 3 Total Penjualan)', fontdict=font_t, pad=20)
    ax.set_xlabel('Periode Tanggal', fontdict=font_l, labelpad=10)
    ax.set_ylabel('Indeks Pertumbuhan (Base 100)', fontdict=font_l, labelpad=10)
    ax.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
    ax.axhline(y=100, color='black', linestyle='-', linewidth=1, alpha=0.5)
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)
    sort_legend(ax)
    plt.tight_layout()
    plt.savefig('rising_star_index.png', bbox_inches='tight')
    print("\nGrafik detail disimpan sebagai 'rising_star_index.png'")

    # --- Chart 2: Nilai Penjualan Asli ---
    fig2, ax2 = plt.subplots(figsize=(15, 8), dpi=100)
    add_top3(ax2, 'total_nilai')
    add_rising(ax2, 'total_nilai')
    ax2.set_title('ANALISIS NILAI PENJUALAN PRODUK RISING STAR\n'
                  '(Nilai Penjualan Asli)', fontdict=font_t, pad=20)
    ax2.set_xlabel('Periode Tanggal', fontdict=font_l, labelpad=10)
    ax2.set_ylabel('Total Nilai Penjualan', fontdict=font_l, labelpad=10)
    ax2.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
    plt.xticks(rotation=45, ha='right', fontsize=10)
    plt.yticks(fontsize=10)
    sort_legend(ax2)
    plt.tight_layout()
    plt.savefig('rising_star_actual.png', bbox_inches='tight')
    print("\nGrafik nilai penjualan disimpan sebagai 'rising_star_actual.png'")

else:
    print("\nTidak ada data untuk di-plot (mungkin tidak ada produk memenuhi kriteria).")